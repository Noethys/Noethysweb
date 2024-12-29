# -*- coding: utf-8 -*-
#  Copyright (c) 2019-2021 Ivan LUCAS.
#  Noethysweb, application de gestion multi-activités.
#  Distribué sous licence GNU GPL.

import json, datetime, time
from django.views.generic import TemplateView
from django.contrib import messages
from django.utils.safestring import mark_safe
from django.http import JsonResponse
from core.views.base import CustomView
from individus.forms.importer_individus import Formulaire_parametres, Formulaire_colonne, DICT_TYPE_DONNEE


def Importer(request):
    time.sleep(1)
    donnees_tableau = json.loads(request.POST["donnees_tableau"])
    colonnes_affectees = json.loads(request.POST["colonnes_affectees"])
    type_import = request.POST["type_import"]
    mode_test = request.POST["mode_test"] == "true"
    masquer_avertissements = request.POST["masquer_avertissements"] == "true"

    # Vérifie qu'il existe des données
    if not donnees_tableau:
        return JsonResponse({"texte_resume": "Votre tableau doit comporter au moins une ligne"}, status=401)
    if not colonnes_affectees:
        return JsonResponse({"texte_resume": "Vous devez affecter au moins une colonne"}, status=401)

    # Initialisation de l'importation
    from individus.utils import utils_importer_individus
    importer = utils_importer_individus.Importer(type_import=type_import, donnees_tableau=donnees_tableau, colonnes_affectees=colonnes_affectees)

    # Recherche d'anomalies et d'avertissements
    liste_anomalies, liste_avertissements = importer.Verifier()
    if liste_anomalies:
        anomalies_donnees = [(anomalie.donnee.num_ligne, anomalie.donnee.num_colonne, anomalie.texte_erreur) for anomalie in liste_anomalies if anomalie.donnee]
        anomalies_lignes = [(anomalie.ligne.num_ligne, anomalie.texte_erreur) for anomalie in liste_anomalies if anomalie.ligne]
        texte_resume = "%d anomalies ont été détectées. Vous pouvez modifier les erreurs directement dans la tableau des données. " % len(liste_anomalies)
        if anomalies_donnees:
            texte_resume += "Survolez les cases oranges dans le tableau pour consulter les erreurs trouvées."
        texte_rapport = texte_resume + "<br><br>" + "<br>".join([anomalie.Get_texte_rapport() for anomalie in liste_anomalies])
        return JsonResponse({"texte_resume": texte_resume, "texte_rapport": texte_rapport, "anomalies_donnees": anomalies_donnees, "anomalies_lignes": anomalies_lignes}, status=401)

    if liste_avertissements and not masquer_avertissements:
        texte_avertissements = "<br>".join(liste_avertissements)
        return JsonResponse({"texte_avertissements": texte_avertissements}, status=401)

    # Si mode test, rapport de succès
    if mode_test:
        return JsonResponse({"succes": True, "rapport_succes": "Aucune anomalie n'a été détectée. Vous pouvez procéder à l'importation."})

    # Enregistrement des données dans la base
    rapport_succes = importer.Enregistrer()
    return JsonResponse({"succes": True, "rapport_succes": rapport_succes})


class View(CustomView, TemplateView):
    menu_code = "importer_individus"
    template_name = "individus/importer_individus.html"

    def get_context_data(self, **kwargs):
        context = super(View, self).get_context_data(**kwargs)
        context["page_titre"] = "Importer des individus"
        if "form_parametres" not in kwargs:
            context["form_parametres"] = Formulaire_parametres(request=self.request)
        return context

    def post(self, request, **kwargs):
        form_parametres = Formulaire_parametres(request.POST, request.FILES, request=self.request)
        if not form_parametres.is_valid():
            return self.render_to_response(self.get_context_data(form_parametres=form_parametres))

        # Ouverture du fichier XLSX
        from io import BytesIO
        from openpyxl import load_workbook

        wb = load_workbook(filename=BytesIO(request.FILES["fichier"].read()), data_only=True, read_only=True)
        ws = wb.active

        def Convert_valeur(valeur):
            # Conversion d'une date anglaise en français
            if valeur and len(valeur) > 7 and valeur[4] == "-" and valeur[7] == "-":
                return datetime.datetime.strptime(valeur,"%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")
            # Si None, renvoie une chaîne vide
            if valeur == "None":
                return ""
            return valeur

        # Lecture du contenu de fichier
        data_excel = []
        ligne_entetes = []
        for num_ligne, row in enumerate(ws.rows):
            ligne = [Convert_valeur(str(cell.value)) for cell in row]
            if num_ligne == 0 and form_parametres.cleaned_data["ligne_entete"]:
                ligne_entetes = ligne
            else:
                data_excel.append(ligne)
        wb.close()

        # Mémorisation des colonnes du fichier Excel
        colonnes = []
        for col in range(0, ws.max_column):
            colonnes.append({"largeur": 180, "entete" : ligne_entetes[col] if form_parametres.cleaned_data["ligne_entete"] else None})

        liste_type_donnee = DICT_TYPE_DONNEE[form_parametres.cleaned_data["type_import"]]

        context = {
            "form_parametres": form_parametres,
            "data_excel": mark_safe(json.dumps(data_excel)),
            "colonnes": colonnes,
            "ligne_entetes": ligne_entetes,
            "type_import": form_parametres.cleaned_data["type_import"],
            "liste_type_donnee": liste_type_donnee[1:],
            "dict_type_donnee": json.dumps({item["code"]: item for item in liste_type_donnee}),
            "form_type_colonne": Formulaire_colonne(request=self.request, type_import=form_parametres.cleaned_data["type_import"]),
        }

        messages.add_message(self.request, messages.INFO, "Le fichier a été lu avec succès. Associez maintenant les colonnes avec les données.")
        return self.render_to_response(self.get_context_data(**context))
